# Data Hall Striping Project
# Tool to generate striping schemes given data hall layout and electrical info
# Wyatt Gardner
# 2025-09-22

import time
from collections import defaultdict
import string
from ortools.sat.python import cp_model
import threading
import shutil
import openpyxl
import sys
import os
BASE_DIR = os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else __file__)
os.chdir(BASE_DIR)

# ==============================================================================
#                                 I. SOLVER SETTINGS
# ==============================================================================
# --- File Configuration ---
INPUT_FILENAME = "data_hall_striping.xlsm"
OUTPUT_FILENAME = "data_hall_striping_OUTPUT.xlsm"
DATA_SHEET_NAME = "Setup & Data"

# --- Solver Behavior ---
UNLIMITED_TIME = True # Lets solver run indefinitely until solver ends or user enters "stop" into terminal
TOTAL_TIME_LIMIT_MINUTES = 30 # Total time for Phase 1 (no split solve) and Phase 2 (split solve) in split mode
PHASE1_BASELINE_SECONDS = 10 # Time to find a starting point for Phase 2
PHASE1_NO_SPLIT_SECONDS = 90 # Total time for the solver in no-split mode
NUM_SEARCH_WORKERS = 16

# --- Objective Weights for Scoring ---
# Any penalty can be set to 0 if minimizing that aspect is not needed.
OVERLOAD_WEIGHT = 20000  # Penalty for each 1 kW of overload
SPLIT_PENALTY_WEIGHT = 100  # Penalty for each circuit that is split (not used if SPLIT_CIRCUITS is false)
DISTANCE_PENALTY_WEIGHT = 0.5 # Penalty for each distance unit a circuit is away from its PDU

# --- Final Report Options ---
PRINT_PDU_SUMMARY = True

# ==============================================================================
#                                 II. EXCEL INPUT
# ==============================================================================
try:
    wb = openpyxl.load_workbook(INPUT_FILENAME, data_only=True)
    ws = wb[DATA_SHEET_NAME]
    
    # Main Operation Mode (from Excel)
    SPLIT_CIRCUITS = ws['D3'].value
    NUM_ROWS = ws['D4'].value
    RACKS_PER_ROW = ws['D6'].value
    NUM_LINEUPS = ws['D7'].value
    PDUS_PER_LINEUP = ws['D8'].value
    NUM_CIRCUITS_PER_PDU = ws['D9'].value
    PDU_CAPACITY = ws['D10'].value
    UPS_CAPACITY = ws['D11'].value

    if any(data is None for data in (SPLIT_CIRCUITS, NUM_ROWS, RACKS_PER_ROW,
        NUM_LINEUPS, PDUS_PER_LINEUP, NUM_CIRCUITS_PER_PDU,
        PDU_CAPACITY, UPS_CAPACITY
    )):
        raise Exception

except FileNotFoundError:
    print(f"FATAL ERROR: Input file not found at '{INPUT_FILENAME}'. Please ensure it is in the same directory.")
    sys.exit(1)
except KeyError:
    print(f"FATAL ERROR: Worksheet '{DATA_SHEET_NAME}' not found in '{INPUT_FILENAME}'.")
    sys.exit(1)
except Exception as e:
    print(f"FATAL ERROR: Could not read setup data from Excel. Check data input. Error: {e}")
    sys.exit(1)

RACK_LOADS = {}
if SPLIT_CIRCUITS:
    loads_flat = []
    # Read all rack unit loads from column Z (26) starting at Z5
    for row in ws.iter_rows(min_row=5, min_col=26, max_col=26, max_row=5 + (NUM_ROWS * RACKS_PER_ROW) - 1):
        loads_flat.append(row[0].value or 0)
    
    # Group the flat list of loads by row
    load_index = 0
    for r in range(1, NUM_ROWS + 1):
        RACK_LOADS[r] = loads_flat[load_index : load_index + RACKS_PER_ROW]
        load_index += RACKS_PER_ROW
else:
    # Read aggregated row loads from column R (18) starting at R5
    row_num = 1
    for row in ws.iter_rows(min_row=5, min_col=18, max_col=18, max_row=5 + NUM_ROWS - 1):
        RACK_LOADS[row_num] = row[0].value or 0
        row_num += 1

# Represents horizontal distance from each PDU to each rack. Rack order is the same as for RACK_LOADS.
PDU_DISTANCES = {
    "A-1":[0,0,1,1,2,2,3,3,4,4,5,5,6,6,7,7,0,0,1,1,2,2,3,3,4,4,5,5,6,6,7,7],
    "A-2":[4,4,3,3,2,2,1,1,0,0,1,1,2,2,3,3,4,4,3,3,2,2,1,1,0,0,1,1,2,2,3,3],
    "A-3":[1,1,0,0,1,1,2,2,3,3,4,4,5,5,6,6,1,1,0,0,1,1,2,2,3,3,4,4,5,5,6,6],
    "B-1":[1,1,0,0,1,1,2,2,3,3,4,4,5,5,6,6,1,1,0,0,1,1,2,2,3,3,4,4,5,5,6,6],
    "B-2":[5,5,4,4,3,3,2,2,1,1,0,0,1,1,2,2,5,5,4,4,3,3,2,2,1,1,0,0,1,1,2,2],
    "B-3":[2,2,1,1,0,0,1,1,2,2,3,3,4,4,5,5,2,2,1,1,0,0,1,1,2,2,3,3,4,4,5,5],
    "C-1":[2,2,1,1,0,0,1,1,2,2,3,3,4,4,5,5,2,2,1,1,0,0,1,1,2,2,3,3,4,4,5,5],
    "C-2":[6,6,5,5,4,4,3,3,2,2,1,1,0,0,1,1,6,6,5,5,4,4,3,3,2,2,1,1,0,0,1,1],
    "C-3":[5,5,4,4,3,3,2,2,1,1,0,0,1,1,2,2,5,5,4,4,3,3,2,2,1,1,0,0,1,1,2,2],
    "D-1":[3,3,2,2,1,1,0,0,1,1,2,2,3,3,4,4,3,3,2,2,1,1,0,0,1,1,2,2,3,3,4,4],
    "D-2":[7,7,6,6,5,5,4,4,3,3,2,2,1,1,0,0,7,7,6,6,5,5,4,4,3,3,2,2,1,1,0,0],
    "D-3":[6,6,5,5,4,4,3,3,2,2,1,1,0,0,1,1,6,6,5,5,4,4,3,3,2,2,1,1,0,0,1,1],
}

# ==============================================================================
#                      III. AUTO-GENERATED SYSTEM VARIABLES
# ==============================================================================
# - These variables are derived from the settings above. Do not edit.
LINEUPS = list(string.ascii_uppercase[:NUM_LINEUPS])
PDUS = [f"{lineup}-{i+1}" for lineup in LINEUPS for i in range(PDUS_PER_LINEUP)]
RACK_IDS = list(range(1, NUM_ROWS + 1))
ALL_CIRCUITS = [(p, c) for p in PDUS for c in range(1, NUM_CIRCUITS_PER_PDU + 1)]

if SPLIT_CIRCUITS:
    RACK_LOADS_AGG = {r: sum(RACK_LOADS[r]) for r in RACK_IDS}
else:
    RACK_LOADS_AGG = RACK_LOADS

# ==============================================================================
#                           CORE LOGIC AND SOLVERS
# ==============================================================================

def stop_listener(solver):
    """Waits for user input to stop the solver."""
    while True:
        command = input()
        if command.lower() == 'stop':
            print("\n...Stop command received. Halting solver gracefully...", flush=True)
            solver.StopSearch()
            break

class SolutionProgressCallback(cp_model.CpSolverSolutionCallback):
    """Prints intermediate solutions and statistics."""

    def __init__(self, variables):
        cp_model.CpSolverSolutionCallback.__init__(self)
        self._vars = variables
        self._start_time = time.time()
        self._last_print_time = self._start_time

    def on_solution_callback(self):
        current_time = time.time()
        if current_time - self._last_print_time > 5.0: # Checkpoint every 5 seconds
            self._last_print_time = current_time
            
            # --- Print progress to console ---
            ol = self.Value(self._vars['total_overload'])
            ol_penalty = ol * OVERLOAD_WEIGHT
            dist = self.Value(self._vars['total_distance'])
            dist_penalty = dist * DISTANCE_PENALTY_WEIGHT
            
            print(f"  > Progress ({current_time - self._start_time:.1f}s):", flush=True)
            print(f"    - Overload: {ol_penalty:>10,} penalty ({ol:,} kW)      ", flush=True)
            print(f"    - Distance: {dist_penalty:>10,} penalty ({dist:,} units)      ", flush=True)

            if 'num_splits' in self._vars:
                splits = self.Value(self._vars['num_splits'])
                split_penalty = splits * SPLIT_PENALTY_WEIGHT
                print(f"    - Splits:   {split_penalty:>10,} penalty ({splits:,} splits)      ", flush=True)

            # --- Reconstruct, normalize, and export solution to Excel ---
            raw_solution = []
            if 'is_unsplit' in self._vars:
                for r in RACK_IDS:
                    row_sol = {'rack': r}
                    if self.Value(self._vars['is_unsplit'][r]):
                        for p, c in ALL_CIRCUITS:
                            if self.Value(self._vars['u_prim'][(r,p,c)]): row_sol['prim'] = (p,c)
                            if self.Value(self._vars['u_sec'][(r,p,c)]): row_sol['sec'] = (p,c)
                    elif self.Value(self._vars['is_prim_split'][r]):
                        row_sol['split_at'] = self.Value(self._vars['split_point'][r])
                        for p, c in ALL_CIRCUITS:
                            if self.Value(self._vars['ps_p1'][(r,p,c)]): row_sol['prim_1'] = (p,c)
                            if self.Value(self._vars['ps_p2'][(r,p,c)]): row_sol['prim_2'] = (p,c)
                            if self.Value(self._vars['ps_s'][(r,p,c)]): row_sol['sec'] = (p,c)
                    elif self.Value(self._vars['is_sec_split'][r]):
                        row_sol['split_at'] = self.Value(self._vars['split_point'][r])
                        for p, c in ALL_CIRCUITS:
                            if self.Value(self._vars['ss_p'][(r,p,c)]): row_sol['prim'] = (p,c)
                            if self.Value(self._vars['ss_s1'][(r,p,c)]): row_sol['sec_1'] = (p,c)
                            if self.Value(self._vars['ss_s2'][(r,p,c)]): row_sol['sec_2'] = (p,c)
                    raw_solution.append(row_sol)
            else:
                for r in RACK_IDS:
                    prim_pdu, prim_c, sec_pdu, sec_c = (None, None, None, None)
                    for p, c in ALL_CIRCUITS:
                        if self.Value(self._vars['primary'][(r,p,c)]): prim_pdu, prim_c = p, c
                        if self.Value(self._vars['secondary'][(r,p,c)]): sec_pdu, sec_c = p, c
                    raw_solution.append({'rack': r, 'prim_pdu': prim_pdu, 'prim_circuit': prim_c, 'sec_pdu': sec_pdu, 'sec_circuit': sec_c})
            
            normalized_solution = normalize_solution(raw_solution)
            generate_excel_output(normalized_solution, is_intermediate=True)

def solve_no_split(time_limit_seconds):
    """Phase 1: Finds the best possible solution without using splits."""
    model = cp_model.CpModel()
    primary, secondary = {}, {}
    for r in RACK_IDS:
        for p, c in ALL_CIRCUITS:
            primary[(r,p,c)] = model.NewBoolVar(f"p_{r}_{p}_{c}")
            secondary[(r,p,c)] = model.NewBoolVar(f"s_{r}_{p}_{c}")

    for r in RACK_IDS:
        model.AddExactlyOne(primary[r,p,c] for p, c in ALL_CIRCUITS)
        model.AddExactlyOne(secondary[r,p,c] for p, c in ALL_CIRCUITS)

    for r in RACK_IDS:
        for lu in LINEUPS:
            pdus_in_lineup = [p for p in PDUS if p.startswith(lu)]
            prim_in_lu = [primary[r,p,c] for p in pdus_in_lineup for c in range(1, NUM_CIRCUITS_PER_PDU + 1)]
            sec_in_lu = [secondary[r,p,c] for p in pdus_in_lineup for c in range(1, NUM_CIRCUITS_PER_PDU + 1)]
            model.Add(sum(prim_in_lu) + sum(sec_in_lu) <= 1)

    for p, c in ALL_CIRCUITS:
        model.Add(sum(primary[r,p,c] + secondary[r,p,c] for r in RACK_IDS) <= 1)

    all_overload_terms = []
    for f_lineup in LINEUPS:
        pdus_in_f = [p for p in PDUS if p.startswith(f_lineup)]
        pdu_loads = {}
        for pdu in PDUS:
            load_terms = []
            for r in RACK_IDS:
                prim_on_pdu_bool = model.NewBoolVar(f"r{r}_prim_on_{pdu}")
                sec_on_pdu_bool = model.NewBoolVar(f"r{r}_sec_on_{pdu}")
                model.Add(sum(primary[r,pdu,c] for c in range(1, NUM_CIRCUITS_PER_PDU + 1)) == 1).OnlyEnforceIf(prim_on_pdu_bool)
                model.Add(sum(primary[r,pdu,c] for c in range(1, NUM_CIRCUITS_PER_PDU + 1)) == 0).OnlyEnforceIf(prim_on_pdu_bool.Not())
                model.Add(sum(secondary[r,pdu,c] for c in range(1, NUM_CIRCUITS_PER_PDU + 1)) == 1).OnlyEnforceIf(sec_on_pdu_bool)
                model.Add(sum(secondary[r,pdu,c] for c in range(1, NUM_CIRCUITS_PER_PDU + 1)) == 0).OnlyEnforceIf(sec_on_pdu_bool.Not())

                is_primary_failed = model.NewBoolVar(f"r{r}_pf_{f_lineup}")
                model.Add(sum(primary[r,p,c] for p in pdus_in_f for c in range(1, NUM_CIRCUITS_PER_PDU + 1)) >= 1).OnlyEnforceIf(is_primary_failed)
                model.Add(sum(primary[r,p,c] for p in pdus_in_f for c in range(1, NUM_CIRCUITS_PER_PDU + 1)) == 0).OnlyEnforceIf(is_primary_failed.Not())

                is_active = model.NewBoolVar(f"r{r}_act_on_{pdu}_f{f_lineup}")
                
                # pa is true iff the primary is on this pdu AND the primary has not failed
                pa = model.NewBoolVar(f"pa_{r}_{pdu}_{f_lineup}")
                model.AddBoolAnd([prim_on_pdu_bool, is_primary_failed.Not()]).OnlyEnforceIf(pa)
                model.AddBoolOr([prim_on_pdu_bool.Not(), is_primary_failed]).OnlyEnforceIf(pa.Not())

                # sa is true iff the secondary is on this pdu AND the primary has failed
                sa = model.NewBoolVar(f"sa_{r}_{pdu}_{f_lineup}")
                model.AddBoolAnd([sec_on_pdu_bool, is_primary_failed]).OnlyEnforceIf(sa)
                model.AddBoolOr([sec_on_pdu_bool.Not(), is_primary_failed.Not()]).OnlyEnforceIf(sa.Not())
                
                model.AddBoolOr([pa, sa]).OnlyEnforceIf(is_active)
                model.AddBoolAnd([pa.Not(), sa.Not()]).OnlyEnforceIf(is_active.Not())

                load_terms.append(RACK_LOADS_AGG[r] * is_active)

            pdu_load = model.NewIntVar(0, 8000, f"load_{pdu}_f{f_lineup}")
            model.Add(pdu_load == sum(load_terms))
            pdu_loads[pdu] = pdu_load
            pdu_overload = model.NewIntVar(0, 8000, f"over_{pdu}_f{f_lineup}")
            model.Add(pdu_overload >= pdu_load - PDU_CAPACITY)
            all_overload_terms.append(pdu_overload)

        for lu in LINEUPS:
            pdus_in_lu = [p for p in PDUS if p.startswith(lu)]
            lineup_load = sum(pdu_loads[p] for p in pdus_in_lu)
            lineup_overload = model.NewIntVar(0, 8000, f"over_lu_{lu}_f{f_lineup}")
            model.Add(lineup_overload >= lineup_load - UPS_CAPACITY)
            all_overload_terms.append(lineup_overload)

    total_overload = model.NewIntVar(0, 8000 * len(LINEUPS) * 4, "total_overload")
    model.Add(total_overload == sum(all_overload_terms))
    
    distance_terms = []
    for r in RACK_IDS:
        for p, c in ALL_CIRCUITS:
            dist = PDU_DISTANCES[p][r-1]
            distance_terms.append(primary[r,p,c] * dist)
            distance_terms.append(secondary[r,p,c] * dist)

    total_distance = model.NewIntVar(0, 99999, "total_distance")
    model.Add(total_distance == sum(distance_terms))
    
    model.Minimize(total_overload * OVERLOAD_WEIGHT + total_distance * DISTANCE_PENALTY_WEIGHT)

    solver = cp_model.CpSolver()
    solver.parameters.num_search_workers = NUM_SEARCH_WORKERS

    if SPLIT_CIRCUITS or not UNLIMITED_TIME:
        solver.parameters.max_time_in_seconds = time_limit_seconds

    callback_vars = {
        'total_overload': total_overload,
        'total_distance': total_distance,
        'primary': primary,
        'secondary': secondary,
    }
    solution_printer = SolutionProgressCallback(callback_vars)
    
    print("\nSolver running. Enter \"stop\" to halt.", flush=True)
    listener_thread = threading.Thread(target=stop_listener, args=(solver,), daemon=True)
    listener_thread.start()
    
    status, results = solver.Solve(model, solution_printer), []

    if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        for r in RACK_IDS:
            prim_pdu, prim_c, sec_pdu, sec_c = (None, None, None, None)
            for p, c in ALL_CIRCUITS:
                if solver.BooleanValue(primary[r,p,c]): prim_pdu, prim_c = p, c
                if solver.BooleanValue(secondary[r,p,c]): sec_pdu, sec_c = p, c
            results.append({'rack': r, 'prim_pdu': prim_pdu, 'prim_circuit': prim_c, 'sec_pdu': sec_pdu, 'sec_circuit': sec_c})
        return results, solver.Value(total_overload)
    else: return [], -1

def solve_with_cp_sat(time_limit_seconds, hint_solution=None):
    """Phase 2: Full-power CP-SAT solver that allows and optimizes splits."""
    model = cp_model.CpModel()

    # --- I. DECISION VARIABLES ---
    is_unsplit = {r: model.NewBoolVar(f'unsplit_{r}') for r in RACK_IDS}
    is_prim_split = {r: model.NewBoolVar(f'prim_split_{r}') for r in RACK_IDS}
    is_sec_split = {r: model.NewBoolVar(f'sec_split_{r}') for r in RACK_IDS}
    split_point = {r: model.NewIntVar(4, RACKS_PER_ROW - 4, f'sp_{r}') for r in RACK_IDS}

    u_prim, u_sec = {}, {}
    ps_p1, ps_p2, ps_s = {}, {}, {}
    ss_p, ss_s1, ss_s2 = {}, {}, {}
    for r in RACK_IDS:
        for p, c in ALL_CIRCUITS:
            u_prim[(r,p,c)], u_sec[(r,p,c)] = model.NewBoolVar(f'u_p_{r}_{p}_{c}'), model.NewBoolVar(f'u_s_{r}_{p}_{c}')
            ps_p1[(r,p,c)], ps_p2[(r,p,c)], ps_s[(r,p,c)] = model.NewBoolVar(f'ps_p1_{r}_{p}_{c}'), model.NewBoolVar(f'ps_p2_{r}_{p}_{c}'), model.NewBoolVar(f'ps_s_{r}_{p}_{c}')
            ss_p[(r,p,c)], ss_s1[(r,p,c)], ss_s2[(r,p,c)] = model.NewBoolVar(f'ss_p_{r}_{p}_{c}'), model.NewBoolVar(f'ss_s1_{r}_{p}_{c}'), model.NewBoolVar(f'ss_s2_{r}_{p}_{c}')

    # --- II. CORE CONSTRAINTS ---
    all_uses_of_circuit = defaultdict(list)
    for r in RACK_IDS:
        model.AddExactlyOne([is_unsplit[r], is_prim_split[r], is_sec_split[r]])
        circuit_vars_r = [u_prim, u_sec, ps_p1, ps_p2, ps_s, ss_p, ss_s1, ss_s2]
        is_active_r = [is_unsplit[r]]*2 + [is_prim_split[r]]*3 + [is_sec_split[r]]*3
        for i, var_group in enumerate(circuit_vars_r):
            circs = [var_group[r,p,c] for p,c in ALL_CIRCUITS]
            model.Add(sum(circs) == 1).OnlyEnforceIf(is_active_r[i])
            model.Add(sum(circs) == 0).OnlyEnforceIf(is_active_r[i].Not())
        for p, c in ALL_CIRCUITS:
              model.Add(ps_p1[r,p,c] + ps_p2[r,p,c] <= 1)
              model.Add(ss_s1[r,p,c] + ss_s2[r,p,c] <= 1)
              for var_group in circuit_vars_r:
                  all_uses_of_circuit[(p, c)].append(var_group[r,p,c])

    for p, c in ALL_CIRCUITS:
        model.Add(sum(all_uses_of_circuit[p, c]) <= 1)

    for r in RACK_IDS:
        for lu in LINEUPS:
            circs_in_lu = [(p,c) for p,c in ALL_CIRCUITS if p.startswith(lu)]
            prim_on_lu = sum(u_prim[r,p,c] + ps_p1[r,p,c] + ps_p2[r,p,c] + ss_p[r,p,c] for p,c in circs_in_lu)
            sec_on_lu = sum(u_sec[r,p,c] + ps_s[r,p,c] + ss_s1[r,p,c] + ss_s2[r,p,c] for p,c in circs_in_lu)
            is_prim_on_lu, is_sec_on_lu = model.NewBoolVar(f'r{r}_p_on_{lu}'), model.NewBoolVar(f'r{r}_s_on_{lu}')
            model.Add(prim_on_lu > 0).OnlyEnforceIf(is_prim_on_lu)
            model.Add(prim_on_lu == 0).OnlyEnforceIf(is_prim_on_lu.Not())
            model.Add(sec_on_lu > 0).OnlyEnforceIf(is_sec_on_lu)
            model.Add(sec_on_lu == 0).OnlyEnforceIf(is_sec_on_lu.Not())
            model.Add(is_prim_on_lu + is_sec_on_lu <= 1)

    # --- III. LOAD & OVERLOAD CALCULATION ---
    prefix_loads = {r: [sum(RACK_LOADS[r][:i]) for i in range(RACKS_PER_ROW + 1)] for r in RACK_IDS}
    load1 = {r: model.NewIntVar(0, 400, f'load1_{r}') for r in RACK_IDS}
    load2 = {r: model.NewIntVar(0, 400, f'load2_{r}') for r in RACK_IDS}
    for r in RACK_IDS:
        model.AddElement(split_point[r], prefix_loads[r], load1[r])
        model.Add(load2[r] == RACK_LOADS_AGG[r] - load1[r])

    all_overload_terms = []
    for f_lineup in LINEUPS:
        circs_in_f = [(p,c) for p,c in ALL_CIRCUITS if p.startswith(f_lineup)]
        pdu_loads = {}
        for pdu in PDUS:
            load_terms_for_pdu = []
            for r in RACK_IDS:
                is_u_prim_failed = model.NewBoolVar(f'fail_up_{r}_{f_lineup}')
                model.Add(sum(u_prim[r,p,c] for p,c in circs_in_f) == 1).OnlyEnforceIf(is_u_prim_failed)
                model.Add(sum(u_prim[r,p,c] for p,c in circs_in_f) == 0).OnlyEnforceIf(is_u_prim_failed.Not())
                is_ps_p1_failed = model.NewBoolVar(f'fail_psp1_{r}_{f_lineup}')
                model.Add(sum(ps_p1[r,p,c] for p,c in circs_in_f) == 1).OnlyEnforceIf(is_ps_p1_failed)
                model.Add(sum(ps_p1[r,p,c] for p,c in circs_in_f) == 0).OnlyEnforceIf(is_ps_p1_failed.Not())
                is_ps_p2_failed = model.NewBoolVar(f'fail_psp2_{r}_{f_lineup}')
                model.Add(sum(ps_p2[r,p,c] for p,c in circs_in_f) == 1).OnlyEnforceIf(is_ps_p2_failed)
                model.Add(sum(ps_p2[r,p,c] for p,c in circs_in_f) == 0).OnlyEnforceIf(is_ps_p2_failed.Not())
                is_ss_p_failed = model.NewBoolVar(f'fail_ssp_{r}_{f_lineup}')
                model.Add(sum(ss_p[r,p,c] for p,c in circs_in_f) == 1).OnlyEnforceIf(is_ss_p_failed)
                model.Add(sum(ss_p[r,p,c] for p,c in circs_in_f) == 0).OnlyEnforceIf(is_ss_p_failed.Not())

                load_from_r_on_this_pdu = model.NewIntVar(0, 400, f'l_{r}_{pdu}_{f_lineup}')
                terms_for_this_row_on_this_pdu = []
                for c_num in range(1, NUM_CIRCUITS_PER_PDU + 1):
                    p, c = pdu, c_num
                    
                    def add_term(base_load, active_bool_def, name):
                        active_bool = model.NewBoolVar(name)
                        model.AddBoolAnd(active_bool_def).OnlyEnforceIf(active_bool)
                        model.AddBoolOr([v.Not() for v in active_bool_def]).OnlyEnforceIf(active_bool.Not())
                        
                        if isinstance(base_load, int):
                           terms_for_this_row_on_this_pdu.append(base_load * active_bool)
                        else: # It's an IntVar
                            term_var = model.NewIntVar(0, 400, f'term_{name}')
                            model.Add(term_var == base_load).OnlyEnforceIf(active_bool)
                            model.Add(term_var == 0).OnlyEnforceIf(active_bool.Not())
                            terms_for_this_row_on_this_pdu.append(term_var)

                    add_term(RACK_LOADS_AGG[r], [is_unsplit[r], u_prim[r,p,c], is_u_prim_failed.Not()], f'act_up_{r}_{p}_{c}')
                    add_term(RACK_LOADS_AGG[r], [is_unsplit[r], u_sec[r,p,c], is_u_prim_failed], f'act_us_{r}_{p}_{c}')
                    add_term(load1[r], [is_prim_split[r], ps_p1[r,p,c], is_ps_p1_failed.Not()], f'act_psp1_{r}_{p}_{c}')
                    add_term(load2[r], [is_prim_split[r], ps_p2[r,p,c], is_ps_p2_failed.Not()], f'act_psp2_{r}_{p}_{c}')
                    add_term(load1[r], [is_prim_split[r], ps_s[r,p,c], is_ps_p1_failed], f'act_pss1_{r}_{p}_{c}')
                    add_term(load2[r], [is_prim_split[r], ps_s[r,p,c], is_ps_p2_failed], f'act_pss2_{r}_{p}_{c}')
                    add_term(RACK_LOADS_AGG[r], [is_sec_split[r], ss_p[r,p,c], is_ss_p_failed.Not()], f'act_ssp_{r}_{p}_{c}')
                    add_term(load1[r], [is_sec_split[r], ss_s1[r,p,c], is_ss_p_failed], f'act_sss1_{r}_{p}_{c}')
                    add_term(load2[r], [is_sec_split[r], ss_s2[r,p,c], is_ss_p_failed], f'act_sss2_{r}_{p}_{c}')

                model.Add(load_from_r_on_this_pdu == sum(terms_for_this_row_on_this_pdu))
                load_terms_for_pdu.append(load_from_r_on_this_pdu)
            
            pdu_load = model.NewIntVar(0, 8000, f"load_{pdu}_f{f_lineup}")
            model.Add(pdu_load == sum(load_terms_for_pdu))
            pdu_loads[pdu] = pdu_load
            pdu_overload = model.NewIntVar(0, 8000, f"over_{pdu}_f{f_lineup}")
            model.Add(pdu_overload >= pdu_load - PDU_CAPACITY)
            all_overload_terms.append(pdu_overload)

        for lu in LINEUPS:
            if lu == f_lineup: continue
            lineup_load = sum(pdu_loads[p] for p in PDUS if p.startswith(lu))
            lineup_overload = model.NewIntVar(0, 8000, f"over_lu_{lu}_f{f_lineup}")
            model.Add(lineup_overload >= lineup_load - UPS_CAPACITY)
            all_overload_terms.append(lineup_overload)

    # --- IV. OBJECTIVE FUNCTION ---
    total_overload = model.NewIntVar(0, 99999, "total_overload")
    model.Add(total_overload == sum(all_overload_terms))
    num_splits = sum(is_prim_split.values()) + sum(is_sec_split.values())

    distance_terms = []
    circuit_vars_all = [u_prim, u_sec, ps_p1, ps_p2, ps_s, ss_p, ss_s1, ss_s2]
    for r in RACK_IDS:
        for p, c in ALL_CIRCUITS:
            dist = PDU_DISTANCES[p][r-1]
            for var_group in circuit_vars_all:
                distance_terms.append(var_group[r,p,c] * dist)

    total_distance = model.NewIntVar(0, 99999, "total_distance")
    model.Add(total_distance == sum(distance_terms))

    model.Minimize(total_overload * OVERLOAD_WEIGHT + 
                   num_splits * SPLIT_PENALTY_WEIGHT +
                   total_distance * DISTANCE_PENALTY_WEIGHT)

    # --- V. SOLVE AND RETURN ---
    solver = cp_model.CpSolver()
    solver.parameters.num_search_workers = NUM_SEARCH_WORKERS
    if not UNLIMITED_TIME:
        solver.parameters.max_time_in_seconds = time_limit_seconds
    
    if hint_solution:
        for r_info in hint_solution:
            r = r_info['rack']
            model.AddHint(is_unsplit[r], 1)
            p_prim, c_prim = r_info['prim']
            p_sec, c_sec = r_info['sec']
            model.AddHint(u_prim[r, p_prim, c_prim], 1)
            model.AddHint(u_sec[r, p_sec, c_sec], 1)

    callback_vars = {
        'total_overload': total_overload,
        'total_distance': total_distance,
        'num_splits': num_splits,
        'is_unsplit': is_unsplit, 'u_prim': u_prim, 'u_sec': u_sec,
        'is_prim_split': is_prim_split, 'ps_p1': ps_p1, 'ps_p2': ps_p2, 'ps_s': ps_s,
        'is_sec_split': is_sec_split, 'ss_p': ss_p, 'ss_s1': ss_s1, 'ss_s2': ss_s2,
        'split_point': split_point
    }
    solution_printer = SolutionProgressCallback(callback_vars)

    print("\nSolver running. Enter \"stop\" to halt.", flush=True)
    listener_thread = threading.Thread(target=stop_listener, args=(solver,), daemon=True)
    listener_thread.start()
    
    status = solver.Solve(model, solution_printer)

    solution = []
    if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        print(f"\nSolver finished with status: {solver.StatusName(status)}")
        for r in RACK_IDS:
            row_sol = {'rack': r}
            if solver.BooleanValue(is_unsplit[r]):
                for p, c in ALL_CIRCUITS:
                    if solver.BooleanValue(u_prim[r,p,c]): row_sol['prim'] = (p,c)
                    if solver.BooleanValue(u_sec[r,p,c]): row_sol['sec'] = (p,c)
            elif solver.BooleanValue(is_prim_split[r]):
                row_sol['split_at'] = solver.Value(split_point[r])
                for p, c in ALL_CIRCUITS:
                    if solver.BooleanValue(ps_p1[r,p,c]): row_sol['prim_1'] = (p,c)
                    if solver.BooleanValue(ps_p2[r,p,c]): row_sol['prim_2'] = (p,c)
                    if solver.BooleanValue(ps_s[r,p,c]): row_sol['sec'] = (p,c)
            elif solver.BooleanValue(is_sec_split[r]):
                row_sol['split_at'] = solver.Value(split_point[r])
                for p, c in ALL_CIRCUITS:
                    if solver.BooleanValue(ss_p[r,p,c]): row_sol['prim'] = (p,c)
                    if solver.BooleanValue(ss_s1[r,p,c]): row_sol['sec_1'] = (p,c)
                    if solver.BooleanValue(ss_s2[r,p,c]): row_sol['sec_2'] = (p,c)
            solution.append(row_sol)
        return solution
    else:
        print(f"\nSolver finished with status: {solver.StatusName(status)}")
        return hint_solution

def get_solution_quality(solution, verbose=False, title="Overload Analysis"):
    rack_map = {row['rack']: row for row in solution}
    if not rack_map: return 0, 0, []
    total_pdu_overload, total_ups_overload, num_splits, total_distance = 0, 0, 0, 0
    
    # Calculate splits and distances
    for r_id, row in rack_map.items():
        if 'split_at' in row:
            num_splits += 1
        
        circuits_used = []
        for key, val in row.items():
            if key in ['prim', 'sec', 'prim_1', 'prim_2', 'sec_1', 'sec_2']:
                circuits_used.append(val)
        
        for pdu, ckt in circuits_used:
            total_distance += PDU_DISTANCES[pdu][r_id-1]

    if verbose: print(f"\n--- {title} ---")

    for f_lineup in LINEUPS:
        if verbose: print(f"\n--- Scenario: Lineup '{f_lineup}' Fails ---")
        pdu_loads, ups_loads = defaultdict(int), defaultdict(int)
        scenario_has_overload = False
        for r in RACK_IDS:
            info = rack_map[r]
            
            if 'prim_1' in info: # Primary is split
                sp, p1, p2, s = info['split_at'], info['prim_1'], info['prim_2'], info['sec']
                load1, load2 = sum(RACK_LOADS[r][:sp]), sum(RACK_LOADS[r][sp:])
                pdu_loads[s[0] if p1[0].startswith(f_lineup) else p1[0]] += load1
                pdu_loads[s[0] if p2[0].startswith(f_lineup) else p2[0]] += load2
            elif 'sec_1' in info: # Secondary is split
                sp, p, s1, s2 = info['split_at'], info['prim'], info['sec_1'], info['sec_2']
                if p[0].startswith(f_lineup):
                    load1, load2 = sum(RACK_LOADS[r][:sp]), sum(RACK_LOADS[r][sp:])
                    pdu_loads[s1[0]] += load1
                    pdu_loads[s2[0]] += load2
                else:
                    pdu_loads[p[0]] += RACK_LOADS_AGG[r]
            else: # Not split
                if 'prim' in info:
                    p, s = info['prim'], info['sec']
                else:
                    p = (info['prim_pdu'], info['prim_circuit'])
                    s = (info['sec_pdu'], info['sec_circuit'])
                active_pdu = s[0] if p[0].startswith(f_lineup) else p[0]
                pdu_loads[active_pdu] += RACK_LOADS_AGG[r]

        for pdu, load in pdu_loads.items():
            ups_loads[pdu[0]] += load
            if load > PDU_CAPACITY:
                total_pdu_overload += load - PDU_CAPACITY
        for load in ups_loads.values():
            if load > UPS_CAPACITY: total_ups_overload += load - UPS_CAPACITY

        if verbose:
            print("  PDU Status:")
            for pdu in sorted(pdu_loads.keys()):
                load = pdu_loads[pdu]
                if load > PDU_CAPACITY:
                    print(f"    - ❌ {pdu}: Load {load} kW > {PDU_CAPACITY} kW (Overload: {load - PDU_CAPACITY} kW)"); scenario_has_overload = True
            print("  UPS Status:")
            for ups in sorted(ups_loads.keys()):
                load = ups_loads[ups]
                if load > UPS_CAPACITY:
                    print(f"    - ❌ UPS {ups}: Load {load} kW > {UPS_CAPACITY} kW (Overload: {load - UPS_CAPACITY} kW)"); scenario_has_overload = True
            if not scenario_has_overload: print("    ✅ No PDU or UPS overloads in this scenario.")

    total_overload = total_pdu_overload + total_ups_overload
    score = (total_overload * OVERLOAD_WEIGHT) + \
            (num_splits * SPLIT_PENALTY_WEIGHT) + \
            (total_distance * DISTANCE_PENALTY_WEIGHT)
    
    if verbose:
        print("\n--- Final Score ---")
        print(f"Overload Penalty: {total_overload * OVERLOAD_WEIGHT:>10,} ({total_overload:,} kW)")
        print(f"Distance Penalty: {total_distance * DISTANCE_PENALTY_WEIGHT:>10,} ({total_distance:,} units)")
        if SPLIT_CIRCUITS:
            print(f"  Split Penalty: {num_splits * SPLIT_PENALTY_WEIGHT:>10,} ({num_splits:,} splits)")
        print("---------------------------------")
        print(f"  TOTAL SCORE:    {score:>10,}")


    return total_overload, score, []

def normalize_solution(raw_solution):
    """Takes the solver's raw output and returns a list of normalized assignments."""
    # 1. Create a detailed list of all assignments with their physical circuit numbers
    assignments = []
    rack_map = {row['rack']: row for row in raw_solution}
    for r in RACK_IDS:
        row_info = rack_map.get(r)
        if not row_info: continue

        # This logic handles all data formats (split, unsplit, phase 1, phase 2)
        # and creates a consistent list of assignments to be normalized.
        is_split_row = 'prim_1' in row_info or 'sec_1' in row_info
        
        if is_split_row:
            for i in range(RACKS_PER_ROW):
                unit_id = i + 1
                entry = {'row': r, 'unit': unit_id}
                if 'prim_1' in row_info:
                    sp, p1, p2, s = row_info['split_at'], row_info['prim_1'], row_info['prim_2'], row_info['sec']
                    prim = p1 if i < sp else p2
                    entry.update({'prim_pdu': prim[0], 'prim_ckt_physical': prim[1], 'sec_pdu': s[0], 'sec_ckt_physical': s[1]})
                else: # 'sec_1' in row_info
                    sp, p, s1, s2 = row_info['split_at'], row_info['prim'], row_info['sec_1'], row_info['sec_2']
                    sec = s1 if i < sp else s2
                    entry.update({'prim_pdu': p[0], 'prim_ckt_physical': p[1], 'sec_pdu': sec[0], 'sec_ckt_physical': sec[1]})
                assignments.append(entry)
        else: # Unsplit row from any solver phase
            if 'prim' in row_info:
                p, s = row_info['prim'], row_info['sec']
            else: # Handles data from solve_no_split
                p = (row_info['prim_pdu'], row_info['prim_circuit'])
                s = (row_info['sec_pdu'], row_info['sec_circuit'])

            if SPLIT_CIRCUITS: # If master mode is split, output should be per-unit
                 for i in range(RACKS_PER_ROW):
                    unit_id = i + 1
                    assignments.append({'row': r, 'unit': unit_id, 'prim_pdu': p[0], 'prim_ckt_physical': p[1], 'sec_pdu': s[0], 'sec_ckt_physical': s[1]})
            else: # If master mode is no-split, output is per-row
                 assignments.append({'row': r, 'unit': 0, 'prim_pdu': p[0], 'prim_ckt_physical': p[1], 'sec_pdu': s[0], 'sec_ckt_physical': s[1]})

    # 2. Find which physical circuits are used by each PDU
    pdu_circuits_used = defaultdict(set)
    for asn in assignments:
        if asn['prim_pdu']: pdu_circuits_used[asn['prim_pdu']].add(asn['prim_ckt_physical'])
        if asn['sec_pdu']: pdu_circuits_used[asn['sec_pdu']].add(asn['sec_ckt_physical'])
    
    # 3. Create the remapping dictionary from physical to normalized circuits
    circuit_remap = {}
    for pdu, physical_circuits in pdu_circuits_used.items():
        sorted_physical_circuits = sorted(list(physical_circuits))
        circuit_remap[pdu] = {physical: i + 1 for i, physical in enumerate(sorted_physical_circuits)}

    # 4. Apply the normalization map to the assignments list
    for asn in assignments:
        if asn['prim_pdu']:
            asn['prim_ckt_norm'] = circuit_remap[asn['prim_pdu']][asn['prim_ckt_physical']]
        if asn['sec_pdu']:
            asn['sec_ckt_norm'] = circuit_remap[asn['sec_pdu']][asn['sec_ckt_physical']]
        
    return assignments

def print_pdu_summary(normalized_solution):
    """Prints a summary of the final circuit assignments using normalized numbers."""
    print("\n--- PDU Circuit Summary (Normalized) ---")
    pdu_map = defaultdict(lambda: defaultdict(list))
    
    for asn in normalized_solution:
        r, u = asn['row'], asn['unit']
        label = f"Row {r}, Unit {u}" if SPLIT_CIRCUITS else f"Row {r}"

        # Group all assignments for a given circuit
        pdu_map[asn['prim_pdu']][asn['prim_ckt_norm']].append(f"{label} (Primary)")
        pdu_map[asn['sec_pdu']][asn['sec_ckt_norm']].append(f"{label} (Secondary)")

    for pdu in PDUS:
        print(f"\n--- PDU: {pdu} ---")
        for slot in range(1, NUM_CIRCUITS_PER_PDU + 1):
            assignments_on_slot = list(set(pdu_map[pdu][slot])) # Use set to show unique assignments
            if assignments_on_slot:
                print(f"  Circuit {slot}: {', '.join(assignments_on_slot)}")
            else:
                print(f"  Circuit {slot}: empty")

def generate_excel_output(normalized_solution, is_intermediate=False):
    """Generates the output Excel file from normalized solution data."""
    if not normalized_solution:
        if not is_intermediate:
            print("WARNING: No solution provided to generate Excel output.")
        return

    if not is_intermediate:
        print(f"\n--- Generating Final Excel Output: {OUTPUT_FILENAME} ---")
    
    try:
        shutil.copy(INPUT_FILENAME, OUTPUT_FILENAME)
        wb = openpyxl.load_workbook(OUTPUT_FILENAME, keep_vba=True)
        ws = wb[DATA_SHEET_NAME]
    except Exception as e:
        if not is_intermediate:
            print(f"❌ Error preparing output Excel file: {e}")
        return

    # Write normalized data to Excel
    start_row = 5
    start_col = 27 if SPLIT_CIRCUITS else 19 # AA or S

    for i, asn in enumerate(normalized_solution):
        output_row_data = [asn.get('prim_pdu'), asn.get('prim_ckt_norm'), asn.get('sec_pdu'), asn.get('sec_ckt_norm')]
        for j, value in enumerate(output_row_data):
            ws.cell(row=start_row + i, column=start_col + j, value=value)

    try:
        wb.save(OUTPUT_FILENAME)
        if not is_intermediate:
            print(f"✅ Successfully created final Excel file: {OUTPUT_FILENAME}")
    except Exception as e:
        if not is_intermediate:
            print(f"❌ Error saving output Excel file: {e}")

if __name__ == "__main__":
    start_time = time.time()
    
    if SPLIT_CIRCUITS:
        print(f"--- Running Phase 1: Generating baseline solution ({PHASE1_BASELINE_SECONDS}s) ---")
        phase1_cp_solution, _ = solve_no_split(PHASE1_BASELINE_SECONDS)
        
        if not phase1_cp_solution:
            print("CRITICAL: CP-SAT Phase 1 failed to find a valid starting solution. Exiting.")
            sys.exit(1)
        
        phase1_solution = [{'rack': row['rack'], 'prim': (row['prim_pdu'], row['prim_circuit']), 'sec': (row['sec_pdu'], row['sec_circuit'])} for row in phase1_cp_solution]
        
        repair_time = TOTAL_TIME_LIMIT_MINUTES * 60 - (time.time() - start_time)
        print(f"\n--- Running Phase 2: CP-SAT Solver with Splits ({'no time limit' if UNLIMITED_TIME else f'{repair_time:.0f}s'}) ---")
        final_raw_solution = solve_with_cp_sat(repair_time, phase1_solution)
        
        if final_raw_solution:
            final_normalized_solution = normalize_solution(final_raw_solution)
            generate_excel_output(final_normalized_solution)
            if PRINT_PDU_SUMMARY:
                print_pdu_summary(final_normalized_solution)
            # Pass raw solution to quality check as it needs the original structure
            get_solution_quality(final_raw_solution, title="Final Solution Analysis", verbose=True)

    else: # NO-SPLIT MODE
        print(f"--- Running in NO-SPLIT mode ({'no time limit' if UNLIMITED_TIME else f'{PHASE1_NO_SPLIT_SECONDS}s'}) ---")
        
        solve_time = PHASE1_NO_SPLIT_SECONDS if not UNLIMITED_TIME else -1
        final_raw_solution, _ = solve_no_split(solve_time)

        if final_raw_solution:
            final_normalized_solution = normalize_solution(final_raw_solution)
            generate_excel_output(final_normalized_solution)
            if PRINT_PDU_SUMMARY:
                print_pdu_summary(final_normalized_solution)
            get_solution_quality(final_raw_solution, title="Final Solution Analysis", verbose=True)
            
    print(f"\nTotal execution time: {time.time() - start_time:.2f} seconds")